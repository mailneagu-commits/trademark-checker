"""
Monitoring router — watch item CRUD, Excel import/template, manual run, history.
"""
import io
import re
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from db import get_db
from monitor_models import WatchItem, SeenTrademark, AlertLog
from monitor_service import run_watch_item

router = APIRouter(prefix="/api/monitor", tags=["monitor"])

# ── Pydantic schemas ──────────────────────────────────────────────────────────

class WatchItemCreate(BaseModel):
    trademark_name:     str
    holder_name:        str  = ""
    nice_classes:       List[str] = []
    offices:            List[str] = ["RO", "EM"]
    notification_email: str
    frequency:          str  = "weekly"   # daily / weekly / monthly


class WatchItemOut(BaseModel):
    id:                 int
    trademark_name:     str
    holder_name:        str
    nice_classes:       List[str]
    offices:            List[str]
    notification_email: str
    frequency:          str
    active:             bool
    created_at:         datetime
    last_checked_at:    Optional[datetime]

    class Config:
        from_attributes = True


class AlertLogOut(BaseModel):
    id:            int
    sent_at:       datetime
    num_new_marks: int
    email_to:      str
    status:        str
    error_msg:     str

    class Config:
        from_attributes = True


# ── CRUD ──────────────────────────────────────────────────────────────────────

@router.get("/watches", response_model=List[WatchItemOut])
def list_watches(db: Session = Depends(get_db)):
    return db.query(WatchItem).order_by(WatchItem.created_at.desc()).all()


@router.post("/watch", response_model=WatchItemOut)
def create_watch(body: WatchItemCreate, db: Session = Depends(get_db)):
    if not body.trademark_name.strip():
        raise HTTPException(400, "Denumirea mărcii este obligatorie.")
    if not body.notification_email.strip():
        raise HTTPException(400, "Email-ul de notificare este obligatoriu.")
    item = WatchItem(**body.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.patch("/watch/{item_id}/toggle", response_model=WatchItemOut)
def toggle_watch(item_id: int, db: Session = Depends(get_db)):
    item = db.get(WatchItem, item_id)
    if not item:
        raise HTTPException(404, "Watch item negăsit.")
    item.active = not item.active
    db.commit()
    db.refresh(item)
    return item


@router.delete("/watch/{item_id}")
def delete_watch(item_id: int, db: Session = Depends(get_db)):
    item = db.get(WatchItem, item_id)
    if not item:
        raise HTTPException(404, "Watch item negăsit.")
    db.query(SeenTrademark).filter(SeenTrademark.watch_item_id == item_id).delete()
    db.query(AlertLog).filter(AlertLog.watch_item_id == item_id).delete()
    db.delete(item)
    db.commit()
    return {"status": "deleted"}


@router.get("/history/{item_id}", response_model=List[AlertLogOut])
def get_history(item_id: int, db: Session = Depends(get_db)):
    item = db.get(WatchItem, item_id)
    if not item:
        raise HTTPException(404, "Watch item negăsit.")
    return (
        db.query(AlertLog)
        .filter(AlertLog.watch_item_id == item_id)
        .order_by(AlertLog.sent_at.desc())
        .limit(50)
        .all()
    )


# ── Manual run ────────────────────────────────────────────────────────────────

@router.post("/watch/{item_id}/run")
async def manual_run(item_id: int, db: Session = Depends(get_db)):
    item = db.get(WatchItem, item_id)
    if not item:
        raise HTTPException(404, "Watch item negăsit.")
    result = await run_watch_item(item, db)
    return result


# ── Excel template ────────────────────────────────────────────────────────────

@router.get("/template")
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Mărci de monitorizat"

    headers = [
        "Denumire Marcă*",
        "Titular",
        "Clase NICE (separate prin virgulă)*",
        "Teritorii (separate prin virgulă)*",
        "Email notificare*",
        "Frecvență (daily/weekly/monthly)",
    ]

    header_fill   = PatternFill("solid", fgColor="1A3C5E")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    col_widths = [30, 30, 35, 30, 35, 30]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell            = ws.cell(row=1, column=col_idx, value=header)
        cell.font       = header_font
        cell.fill       = header_fill
        cell.alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border     = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 36

    # Example row
    example = ["ACME", "ACME România SRL", "35, 42", "RO, EM", "office@firma.ro", "weekly"]
    example_fill = PatternFill("solid", fgColor="EBF5FB")
    for col_idx, val in enumerate(example, start=1):
        cell           = ws.cell(row=2, column=col_idx, value=val)
        cell.fill      = example_fill
        cell.border    = thin_border
        cell.alignment = Alignment(vertical="center")

    # Note row
    ws.cell(row=3, column=1, value="* câmpuri obligatorii")
    ws.cell(row=3, column=1).font = Font(italic=True, color="888888")
    ws.cell(row=4, column=1, value="Teritorii acceptate: RO, EM, EU, DE, FR, IT, ES, UK, US, WO (sau orice cod de țară din TMview)")
    ws.cell(row=4, column=1).font = Font(italic=True, color="888888")
    ws.merge_cells("A4:F4")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="template_monitorizare_marci.xlsx"'},
    )


# ── Excel import ──────────────────────────────────────────────────────────────

def _parse_classes(raw: str) -> List[str]:
    return [c.strip() for c in re.split(r"[,;\s]+", str(raw)) if c.strip().isdigit()]


def _parse_offices(raw: str) -> List[str]:
    return [o.strip().upper() for o in re.split(r"[,;\s]+", str(raw)) if o.strip()]


def _parse_frequency(raw: str) -> str:
    val = str(raw).strip().lower()
    return val if val in ("daily", "weekly", "monthly") else "weekly"


@router.get("/bulletin-status")
def bulletin_status():
    """Returnează lista tuturor buletinelor descărcate (OSIM + EUIPO)."""
    from scrapers.osim_bulletin  import list_processed as osim_list
    from scrapers.euipo_bulletin import list_processed as euipo_list
    return {
        "osim":  osim_list(),
        "euipo": euipo_list(),
    }


@router.get("/bulletin-marks")
def get_bulletin_marks(source: str, date: str):
    """
    Returnează mărcile dintr-un buletin descărcat anterior.
    source: 'osim' | 'euipo'
    date: 'YYYY-MM-DD'
    """
    from datetime import date as date_type

    try:
        td = date_type.fromisoformat(date)
    except ValueError:
        raise HTTPException(400, f"Dată invalidă: {date}")

    if source == "osim":
        from scrapers.osim_bulletin import _prev_working_day, _date_slug, _parse_pdf, CACHE_DIR
        import os
        working = _prev_working_day(td)
        slug    = _date_slug(working)
        pdf     = os.path.join(CACHE_DIR, f"{slug}.pdf")
        if not os.path.exists(pdf):
            raise HTTPException(404, "Buletinul OSIM pentru această dată nu a fost descărcat încă.")
        marks = _parse_pdf(pdf)

    elif source == "euipo":
        from scrapers.euipo_bulletin import _prev_working_day, _date_slug, _parse_bulletin_xml, CACHE_DIR
        import os
        working = _prev_working_day(td)
        slug    = _date_slug(working)
        xml     = os.path.join(CACHE_DIR, f"{slug}.xml")
        if not os.path.exists(xml):
            raise HTTPException(404, "Buletinul EUIPO pentru această dată nu a fost descărcat încă.")
        marks = _parse_bulletin_xml(xml)

    else:
        raise HTTPException(400, "source trebuie să fie 'osim' sau 'euipo'")

    return {"source": source, "date": date, "total": len(marks), "marks": marks}


@router.post("/bulletin-fetch")
async def trigger_bulletin_fetch(
    source: str = "both",          # "osim" | "euipo" | "both"
    target_date: Optional[str] = None,   # ISO date string "YYYY-MM-DD"
):
    """
    Descarcă buletinul pentru o dată specificată (sau cel mai recent dacă lipsește).
    source: "osim" | "euipo" | "both"
    target_date: "YYYY-MM-DD" (opțional, implicit azi)
    """
    import asyncio
    from datetime import date as date_type
    from scrapers.osim_bulletin  import fetch_osim_for_date,  fetch_latest_osim
    from scrapers.euipo_bulletin import fetch_euipo_for_date, fetch_latest_euipo

    if target_date:
        try:
            td = date_type.fromisoformat(target_date)
        except ValueError:
            raise HTTPException(400, f"Dată invalidă: {target_date}. Folosiți formatul YYYY-MM-DD.")
    else:
        td = None

    loop = asyncio.get_event_loop()
    result: dict = {"source": source, "target_date": target_date or "latest"}

    if source in ("osim", "both"):
        if td:
            marks, info = await loop.run_in_executor(None, fetch_osim_for_date, td)
        else:
            marks = await loop.run_in_executor(None, fetch_latest_osim)
            info  = {}
        result["osim"] = {"marks": len(marks), **info}

    if source in ("euipo", "both"):
        if td:
            marks, info = await loop.run_in_executor(None, fetch_euipo_for_date, td)
        else:
            marks = await loop.run_in_executor(None, fetch_latest_euipo)
            info  = {}
        result["euipo"] = {"marks": len(marks), **info}

    return result


@router.post("/import")
def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Fișierul trebuie să fie .xlsx sau .xls")

    from openpyxl import load_workbook

    content = file.file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Fișier Excel invalid: {e}")

    ws = wb.active
    imported  = []
    skipped   = []
    errors    = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v is None for v in row):
            continue

        # Skip note rows (first cell starts with "*" or is italic note)
        first = str(row[0] or "").strip()
        if first.startswith("*") or first.startswith("Teritorii") or first.startswith("câmp"):
            continue

        trademark_name     = str(row[0] or "").strip() if len(row) > 0 else ""
        holder_name        = str(row[1] or "").strip() if len(row) > 1 else ""
        nice_classes_raw   = str(row[2] or "").strip() if len(row) > 2 else ""
        offices_raw        = str(row[3] or "").strip() if len(row) > 3 else ""
        notification_email = str(row[4] or "").strip() if len(row) > 4 else ""
        frequency_raw      = str(row[5] or "").strip() if len(row) > 5 else "weekly"

        if not trademark_name:
            skipped.append({"row": row_idx, "reason": "Denumire marcă lipsă"})
            continue
        if not notification_email or "@" not in notification_email:
            errors.append({"row": row_idx, "trademark": trademark_name, "reason": "Email invalid sau lipsă"})
            continue

        nice_classes = _parse_classes(nice_classes_raw)
        if not nice_classes:
            errors.append({"row": row_idx, "trademark": trademark_name, "reason": "Clase NICE invalide sau lipsă"})
            continue

        offices   = _parse_offices(offices_raw) or ["RO", "EM"]
        frequency = _parse_frequency(frequency_raw)

        item = WatchItem(
            trademark_name     = trademark_name,
            holder_name        = holder_name,
            nice_classes       = nice_classes,
            offices            = offices,
            notification_email = notification_email,
            frequency          = frequency,
        )
        db.add(item)
        imported.append({"row": row_idx, "trademark": trademark_name, "email": notification_email})

    db.commit()

    return {
        "imported": len(imported),
        "skipped":  len(skipped),
        "errors":   len(errors),
        "details": {
            "imported": imported,
            "skipped":  skipped,
            "errors":   errors,
        },
    }
