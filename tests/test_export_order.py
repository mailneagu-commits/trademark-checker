import io
import unittest

from openpyxl import load_workbook
from docx import Document

from backend.export import build_excel, build_word


class ExportOrderTests(unittest.TestCase):
    def test_excel_orders_by_office_then_filing_date(self):
        results = [
            {
                "tmName": "WO MID",
                "office": "WO",
                "officeName": "WIPO",
                "status": "registered",
                "applicationDate": "2019-09-09",
                "registrationDate": "2020-01-01",
                "applicants": [{"name": "Owner WO"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 99},
            },
            {
                "tmName": "RO NEW",
                "office": "RO",
                "officeName": "OSIM Romania",
                "status": "registered",
                "applicationDate": "2023-06-10",
                "registrationDate": "2024-01-01",
                "applicants": [{"name": "Owner RO New"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 60},
            },
            {
                "tmName": "EM OLD",
                "office": "EM",
                "officeName": "EUIPO",
                "status": "registered",
                "applicationDate": "2017-04-20",
                "registrationDate": "2018-01-01",
                "applicants": [{"name": "Owner EM Old"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 95},
            },
            {
                "tmName": "RO OLD",
                "office": "RO",
                "officeName": "OSIM Romania",
                "status": "registered",
                "applicationDate": "2018-02-01",
                "registrationDate": "2025-01-01",
                "applicants": [{"name": "Owner RO Old"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 55},
            },
        ]

        data = build_excel("ALFA", ["35"], ["RO"], results, [], [], [])
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Raport Similaritate"]

        self.assertEqual(ws["A5"].value, "Mărci active (4)")
        names = [ws[f"E{row}"].value for row in range(5, 9)]
        self.assertEqual([ws[f"E{row}"].value for row in range(6, 10)], ["RO OLD", "RO NEW", "EM OLD", "WO MID"])

    def test_excel_keeps_primary_territory_before_em_and_other_selected_territories(self):
        results = [
            {
                "tmName": "DE OLD",
                "office": "DE",
                "officeName": "DPMA",
                "status": "registered",
                "applicationDate": "2016-01-01",
                "registrationDate": "2017-01-01",
                "applicants": [{"name": "Owner DE"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 95},
            },
            {
                "tmName": "EM OLD",
                "office": "EM",
                "officeName": "EUIPO",
                "status": "registered",
                "applicationDate": "2015-01-01",
                "registrationDate": "2016-01-01",
                "applicants": [{"name": "Owner EM"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 99},
            },
            {
                "tmName": "RO OLD",
                "office": "RO",
                "officeName": "OSIM Romania",
                "status": "registered",
                "applicationDate": "2018-02-01",
                "registrationDate": "2019-01-01",
                "applicants": [{"name": "Owner RO"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 50},
            },
            {
                "tmName": "WO OLD",
                "office": "WO",
                "officeName": "WIPO",
                "status": "registered",
                "applicationDate": "2014-01-01",
                "registrationDate": "2015-01-01",
                "applicants": [{"name": "Owner WO"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 97},
            },
        ]

        data = build_excel("ALFA", ["35"], ["RO", "DE", "EM"], results, [], [], [])
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Raport Similaritate"]

        self.assertEqual([ws[f"E{row}"].value for row in range(6, 10)], ["RO OLD", "EM OLD", "WO OLD", "DE OLD"])

    def test_excel_conclusions_include_expired_marks_in_risk_counts(self):
        active_results = [
            {
                "tmName": "ACTIVE MEDIUM",
                "office": "RO",
                "officeName": "OSIM Romania",
                "status": "registered",
                "applicationDate": "2020-01-01",
                "registrationDate": "2021-01-01",
                "applicants": [{"name": "Owner Active"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 67},
            },
        ]
        expired_results = [
            {
                "tmName": "EXPIRED HIGH",
                "office": "RO",
                "officeName": "OSIM Romania",
                "status": "expired",
                "applicationDate": "2010-01-01",
                "registrationDate": "2011-01-01",
                "applicants": [{"name": "Owner Expired"}],
                "niceClass": ["35"],
                "goodAndServices": [],
                "similarity": {"combined_score": 82},
            },
        ]

        data = build_excel("ALFA", ["35"], ["RO"], active_results, [], expired_results, [])
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Concluzii"]

        self.assertEqual(ws["B5"].value, 1)
        self.assertEqual(ws["B6"].value, 1)
        self.assertEqual(ws["B8"].value, 1)

    def test_word_places_expired_chapter_after_conclusions(self):
        active_results = [{
            "tmName": "ACTIVE MARK",
            "office": "RO",
            "officeName": "OSIM Romania",
            "status": "registered",
            "applicationDate": "2020-01-01",
            "registrationDate": "2021-01-01",
            "applicants": [{"name": "Active Owner"}],
            "niceClass": ["35"],
            "niceDetailed": [{"class": "35", "short": "publicitate"}],
            "goodAndServices": [{"niceClass": "35", "goodsAndServices": "servicii active"}],
            "similarity": {"combined_score": 67, "textual_score": 66, "phonetic_score": 60, "jaro_winkler": 72, "levenshtein_distance": 4},
        }]
        expired_results = [{
            "tmName": "EXPIRED MARK",
            "office": "RO",
            "officeName": "OSIM Romania",
            "status": "expired",
            "applicationDate": "2010-01-01",
            "registrationDate": "2011-01-01",
            "applicants": [{"name": "Expired Owner"}],
            "niceClass": ["35"],
            "niceDetailed": [{"class": "35", "short": "publicitate"}],
            "goodAndServices": [{"niceClass": "35", "goodsAndServices": "servicii expirate"}],
            "similarity": {"combined_score": 82, "textual_score": 80, "phonetic_score": 70, "jaro_winkler": 88, "levenshtein_distance": 2},
        }]

        data = build_word("TEST", ["35"], ["RO"], active_results, [], expired_results, [])
        doc = Document(io.BytesIO(data))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())

        self.assertIn("Mărci active (1)", text)
        self.assertIn("Concluzii și recomandări", text)
        self.assertIn("Mărci expirate (1)", text)
        self.assertLess(text.index("Concluzii și recomandări"), text.index("Mărci expirate (1)"))


if __name__ == "__main__":
    unittest.main()
